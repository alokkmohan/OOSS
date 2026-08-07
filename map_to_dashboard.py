"""
Map classified records into the "OOS Dashboard 75Districts" template.

The dashboard template ships with live formulas (COUNTIF/COUNTIFS) on the
'District Summary (75)', 'Reason-wise Analysis', 'Gender & Category', and
'Class-wise Analysis' sheets that all read from 'Master Raw Data'. This
script only needs to fill 'Master Raw Data' in the exact column order/values
those formulas expect — everything else recalculates automatically when the
file is opened in Excel (or run through LibreOffice --convert-to, see below).

USAGE
-----
    python map_to_dashboard.py CLASSIFIED.pkl DASHBOARD_TEMPLATE.xlsx -o OUTPUT.xlsx

Then, to get live numbers without opening Excel by hand:
    soffice --headless --convert-to xlsx --outdir out/ OUTPUT.xlsx
(This forces LibreOffice to recalculate all formulas and re-save.)
"""
import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.cell.cell import MergedCell

# District names in the source data don't always match the dashboard's
# canonical 75-district spelling. Extend this as new mismatches show up.
DISTRICT_ALIAS = {
    "GB NAGAR": "Gautam Buddha Nagar",
    "HAMIRPUR (U.P.)": "Hamirpur",
    "KHERI": "Lakhimpur Kheri",
    "LAKHIMPUR KHERI": "Lakhimpur Kheri",
    "SHRAWASTI": "Shravasti",
    "BHADOI": "Bhadohi",
}

NOT_STUDYING_REASON_PATTERNS = [
    ("Marriage", r"marri|शादी|ससुराल"),
    ("Migrated for Labour/Work", r"labour|labor|labure|मजदूरी|working|\bjob\b|naukri|migrat"),
    ("Admission Not Taken / TC Issue", r"tc\s*issu|no\s*admi|not\s*admi|admisson"),
    ("Household / Domestic Responsibility", r"घरेलू|कृषि|किसानी|दुकान|home\s*work|household"),
    ("Health / Medical Reason", r"sick|health|ill\b|disab"),
    ("Financial Problem", r"poor\s*econ|economic|financial"),
    ("Refused to Continue Studies",
     r"not\s*intere?st|not\s*want\s*to\s*stud|unwilling|refused|नहीं\s*पढ़ना|"
     r"guardian\s*not\s*inter|parents?\s*not\s*inter|अभिभावक"),
]
UNCLEAR_REASON_PATTERNS = [
    ("Wrong / Invalid Number", r"wrong\s*no|wrong\s*number|wrong\s*phone|invalid\s*number|incorrect\s*phone"),
    ("Switched Off / Not Reachable", r"switch\s*off|switched\s*off|swich\s*off|swtich\s*off"),
    ("Call Not Received", r"not\s*receiv|no\s*answer|not\s*answer|call\s*not|no\s*respond|not\s*respond"),
    ("No Information Available", r"no\s*info|data\s*not|active\s*for\s*import|status\s*not\s*known|no\s*data"),
    ("Refused to Share Information", r"refused\s*to\s*share|not\s*share"),
]


def norm_district(d):
    du = str(d).strip().upper()
    if du in DISTRICT_ALIAS:
        return DISTRICT_ALIAS[du]
    return str(d).strip().title()


def map_status_category(row):
    if row.get("Death_Flag") == "Death Reported":
        return "Death Case"
    sc = row["Status_Category"]
    if sc == "Known - Studying":
        return "Known - Studying"
    if sc == "Known - Not Studying / Dropout":
        return "Known - Not Studying / Dropout"
    return "Unclear"


def map_current_status_simple(scd):
    return {
        "Death Case": "Deceased",
        "Known - Studying": "Studying",
        "Known - Not Studying / Dropout": "Not Studying",
    }.get(scd, "Unclear")


def detect_reason(row):
    scd = row["__status_cat_dash"]
    text = f"{row['current Status']} {row['Remark']}"
    if scd == "Death Case":
        return "Death (Student/Family Member)"
    if scd == "Known - Not Studying / Dropout":
        for name, pat in NOT_STUDYING_REASON_PATTERNS:
            if re.search(pat, str(text), re.IGNORECASE):
                return name
        return "Other"
    if scd == "Unclear":
        for name, pat in UNCLEAR_REASON_PATTERNS:
            if re.search(pat, str(text), re.IGNORECASE):
                return name
        if str(row["current Status"]).strip() in ("", "nan", "None") and str(row["Remark"]).strip() in ("", "nan", "None"):
            return "Yet to be Contacted"
        return "Other"
    return ""


def map_willing(row):
    if row["__status_cat_dash"] != "Known - Not Studying / Dropout":
        return ""
    return "No" if row.get("Willingness_Category") == "Unwilling - Does Not Want to Study" else "Yes"


def build_master_raw_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["__district_norm"] = df["District Name"].apply(norm_district)
    df["__status_cat_dash"] = df.apply(map_status_category, axis=1)
    df["__current_status_simple"] = df["__status_cat_dash"].apply(map_current_status_simple)
    df["__reason_category"] = df.apply(detect_reason, axis=1)
    df["__willing_resume"] = df.apply(map_willing, axis=1)
    df["__remark_verbatim"] = (
        df["current Status"].astype(str).str.strip() + " | " + df["Remark"].astype(str).str.strip()
    )

    return pd.DataFrame({
        "Sr No": range(1, len(df) + 1),
        "District Name": df["__district_norm"],
        "Block Name": df["Block Name"],
        "Student Name": df["Student Name"],
        "Father Name": df["Father Name"],
        "Gender": df["Gender"],
        "Social Category": df["Category"],
        "Class": df["Class"],
        "Dropout Year": df["Droupout Year"],
        "Address": df["Address"],
        "Mobile No.": df["Mobile No."],
        "Current Status (Study/Not Study)": df["__current_status_simple"],
        "Status Category": df["__status_cat_dash"],
        "Reason Category": df["__reason_category"],
        "Remark (Verbatim)": df["__remark_verbatim"],
        "Willing to Resume Studies": df["__willing_resume"],
        "Referred to Open Schooling (NIOS/UPSOSB)": "",
        "Duplicate PEN/Record Flag": "",
        "Data Collected By": "",
        "Collection Date": "",
        "Interested in Studying via NIOS (Yes/No)": "",
    })


def widen_formula_ranges(ws, old_bound="30000", new_bound="50000"):
    """The template's formulas are hardcoded to a 30,000-row range. If your
    dataset is bigger than that, formulas silently miss the extra rows."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=") and old_bound in cell.value:
                cell.value = cell.value.replace(old_bound, new_bound)
                count += 1
    return count


def write_master_raw_data(wb, out_df: pd.DataFrame, start_row=4):
    ws = wb["Master Raw Data"]
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, start_row)):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    max_row_needed = start_row + len(out_df) - 1
    widened = widen_formula_ranges(wb["District Summary (75)"])
    widened += widen_formula_ranges(wb["Reason-wise Analysis"])
    widened += widen_formula_ranges(wb["Gender & Category"])
    widened += widen_formula_ranges(wb["Class-wise Analysis"])
    print(f"Widened {widened} formula ranges to cover {max_row_needed} rows.")

    for i, row_vals in enumerate(out_df.values.tolist(), start=start_row):
        for col_idx, val in enumerate(row_vals, start=1):
            if pd.isna(val):
                val = ""
            ws.cell(row=i, column=col_idx, value=val)


def main():
    parser = argparse.ArgumentParser(description="Populate the OOS Dashboard template from classified data.")
    parser.add_argument("classified_pkl", type=Path, help="Pickle produced by classify.py (e.g. output_classified.pkl)")
    parser.add_argument("template_xlsx", type=Path, help="Path to the blank OOS_Dashboard_75Districts.xlsx template")
    parser.add_argument("-o", "--output", type=Path, required=True, help="Output populated .xlsx path")
    args = parser.parse_args()

    df = pd.read_pickle(args.classified_pkl)
    out_df = build_master_raw_data(df)
    print("Status Category breakdown:")
    print(out_df["Status Category"].value_counts())

    wb = openpyxl.load_workbook(args.template_xlsx)
    write_master_raw_data(wb, out_df)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Saved: {args.output}")
    print(
        "NOTE: formulas won't show computed values until the file is opened in Excel, "
        "or recalculated headlessly, e.g.:\n"
        f"  soffice --headless --convert-to xlsx --outdir <dir> {args.output}"
    )


if __name__ == "__main__":
    main()
